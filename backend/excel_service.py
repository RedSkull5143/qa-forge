import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel_from_test_cases(test_cases: list) -> bytes:
    """
    Converts a list of test case Pydantic models (or dicts) into a
    professionally formatted Excel report in memory.

    Features:
        - Styled header row with dark blue background and white bold text
        - Alternating row colors for readability
        - Auto-sized column widths
        - Text wrapping for multi-line cells (Steps, Expected Results)
        - Thin borders on every cell
        - Color-coded Type column (Positive=Green, Negative=Red, Edge=Orange)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # ── Define Styles ──────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Alternating row fills
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Type column color coding
    type_colors = {
        "Positive": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Negative": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Edge":     PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid"),
    }

    # ── Column Configuration ───────────────────────────────────────
    columns = [
        {"header": "TC ID",             "key": "id",               "width": 10},
        {"header": "Title",             "key": "title",            "width": 35},
        {"header": "Type",              "key": "type",             "width": 12},
        {"header": "Summary",           "key": "summary",          "width": 45},
        {"header": "Preconditions",     "key": "preconditions",    "width": 40},
        {"header": "Steps",             "key": "steps",            "width": 50},
        {"header": "Expected Results",  "key": "expected_results", "width": 45},
        {"header": "Priority",          "key": "priority",         "width": 12},
    ]

    # ── Write Header Row ───────────────────────────────────────────
    for col_idx, col_cfg in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_cfg["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    for col_idx, col_cfg in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_cfg["width"]

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    # ── Write Data Rows ────────────────────────────────────────────
    for row_idx, tc in enumerate(test_cases, start=2):
        # Support both Pydantic models and plain dicts
        tc_data = tc.model_dump() if hasattr(tc, "model_dump") else tc
        row_fill = light_fill if row_idx % 2 == 0 else white_fill

        for col_idx, col_cfg in enumerate(columns, start=1):
            raw_value = tc_data.get(col_cfg["key"], "")

            # Convert lists to numbered/bulleted strings
            if isinstance(raw_value, list):
                if col_cfg["key"] == "steps":
                    raw_value = "\n".join(f"{i+1}. {s}" for i, s in enumerate(raw_value))
                else:
                    raw_value = "\n".join(f"• {item}" for item in raw_value)

            cell = ws.cell(row=row_idx, column=col_idx, value=str(raw_value))
            cell.font = data_font
            cell.border = thin_border

            # Center-align short columns (ID, Type, Priority)
            if col_cfg["key"] in ("id", "type", "priority"):
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment

            # Apply alternating row color
            cell.fill = row_fill

            # Override fill for the Type column with color coding
            if col_cfg["key"] == "type":
                cell.fill = type_colors.get(str(raw_value), row_fill)
                cell.font = Font(name="Calibri", size=11, bold=True)

    # ── Auto-fit Row Heights (approximate) ─────────────────────────
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = 1
        for cell in row:
            if cell.value:
                lines = str(cell.value).count("\n") + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

    # ── Write to Memory Buffer ─────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
